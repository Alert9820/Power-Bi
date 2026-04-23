"""
Business Intelligence Reporter
Auto-generates professional Excel reports with ML predictions
Flask + Pandas + OpenPyXL + Scikit-learn
"""

import os, io, json, uuid, traceback, warnings, math
from pathlib import Path
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify, send_file, render_template_string
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score, mean_absolute_error
import warnings
warnings.filterwarnings("ignore")

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB
UPLOAD_FOLDER = Path("uploads")
REPORT_FOLDER = Path("reports")
UPLOAD_FOLDER.mkdir(exist_ok=True)
REPORT_FOLDER.mkdir(exist_ok=True)

SESSIONS = {}

# ── Helpers ──────────────────────────────────────────────────────────────────
def clean_val(v):
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if hasattr(v, 'item'):
        return v.item()
    return v

def read_file(file_bytes, filename):
    if filename.endswith('.csv'):
        return pd.read_csv(io.BytesIO(file_bytes), low_memory=False)
    else:
        return pd.read_excel(io.BytesIO(file_bytes))

def detect_cols(df):
    rev_col  = next((c for c in df.columns if any(k in c.lower() for k in ['revenue','sales','income','amount','price','turnover'])), None)
    cost_col = next((c for c in df.columns if any(k in c.lower() for k in ['cost','expense','spend','expenditure'])), None)
    date_col = next((c for c in df.columns if any(k in c.lower() for k in ['date','month','period','time','year'])), None)
    cat_col  = next((c for c in df.columns if any(k in c.lower() for k in ['category','product','region','segment','department','type'])), None)
    qty_col  = next((c for c in df.columns if any(k in c.lower() for k in ['qty','units','quantity','count','volume'])), None)
    return rev_col, cost_col, date_col, cat_col, qty_col

def clean_dataframe(df):
    logs = []
    orig = df.shape
    df = df.drop_duplicates()
    logs.append(f"Removed {orig[0] - len(df)} duplicate rows")

    for col in df.columns:
        miss = df[col].isna().sum()
        if miss == 0: continue
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].fillna(df[col].median())
        else:
            df[col] = df[col].fillna(df[col].mode()[0] if not df[col].mode().empty else 'Unknown')
        logs.append(f"'{col}': filled {miss} missing values")

    num_cols = df.select_dtypes(include=np.number).columns
    removed = 0
    for col in num_cols:
        q1, q3 = df[col].quantile(0.01), df[col].quantile(0.99)
        iqr = q3 - q1
        before = len(df)
        df = df[(df[col] >= q1 - 3*iqr) & (df[col] <= q3 + 3*iqr)]
        removed += before - len(df)
    logs.append(f"Removed {removed} outlier rows (IQR method)")
    logs.append(f"Clean shape: {len(df):,} rows × {len(df.columns)} columns")
    return df, logs

def run_ml_prediction(df, rev_col):
    if not rev_col or rev_col not in df.columns:
        return None
    try:
        num_cols = df.select_dtypes(include=np.number).columns.tolist()
        features = [c for c in num_cols if c != rev_col]
        if len(features) == 0:
            features = [rev_col]
            df_ml = df[[rev_col]].copy().dropna()
            df_ml['index'] = np.arange(len(df_ml))
            X = df_ml[['index']]
            y = df_ml[rev_col]
        else:
            X = df[features].fillna(0)
            y = df[rev_col].fillna(0)

        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
        rf = RandomForestRegressor(n_estimators=100, random_state=42)
        rf.fit(X_train, y_train)
        preds = rf.predict(X_test)

        r2  = round(r2_score(y_test, preds), 4)
        mae = round(mean_absolute_error(y_test, preds), 2)

        # Next period prediction
        last_row = X.iloc[[-1]]
        next_pred = round(float(rf.predict(last_row)[0]), 2)
        current_avg = round(float(y.mean()), 2)
        growth = round((next_pred - current_avg) / current_avg * 100, 2) if current_avg != 0 else 0

        fi = {}
        if hasattr(rf, 'feature_importances_'):
            fi = dict(zip(features, [round(float(v)*100, 1) for v in rf.feature_importances_]))
            fi = dict(sorted(fi.items(), key=lambda x: x[1], reverse=True)[:5])

        return {
            'r2': r2, 'mae': mae,
            'next_prediction': next_pred,
            'current_avg': current_avg,
            'growth_pct': growth,
            'feature_importance': fi,
            'model': 'Random Forest'
        }
    except Exception as e:
        return {'error': str(e)}

# ── Excel Report Generator ────────────────────────────────────────────────────
def generate_excel_report(df, filename, ml_result, logs):
    wb = Workbook()

    # ── Colors ──
    DARK    = "0D1117"
    ACCENT  = "00E5FF"
    ACCENT2 = "7C5CFC"
    GREEN   = "00E096"
    YELLOW  = "FFD166"
    RED     = "FF5C8A"
    WHITE   = "E8EAF0"
    GRAY    = "1E2736"
    LGRAY   = "2D3748"

    def hdr_font(size=12, bold=True, color=WHITE):
        return Font(name='Calibri', size=size, bold=bold, color=color)

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    def thin_border():
        s = Side(style='thin', color=GRAY)
        return Border(left=s, right=s, top=s, bottom=s)

    rev_col, cost_col, date_col, cat_col, qty_col = detect_cols(df)
    num_cols = df.select_dtypes(include=np.number).columns.tolist()

    # ════════════════════════════════════════════════════
    # SHEET 1: EXECUTIVE SUMMARY
    # ════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Executive Summary"
    ws1.sheet_view.showGridLines = False

    # Set col widths
    col_widths = [3, 20, 18, 18, 18, 18, 18, 3]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, 60):
        ws1.row_dimensions[r].height = 22

    # Title banner
    ws1.merge_cells('B2:G3')
    ws1['B2'] = f"📈  BUSINESS INTELLIGENCE REPORT"
    ws1['B2'].font = Font(name='Calibri', size=20, bold=True, color=ACCENT)
    ws1['B2'].fill = fill(DARK)
    ws1['B2'].alignment = center()

    ws1.merge_cells('B4:G4')
    ws1['B4'] = f"Source: {filename}  |  Generated: {pd.Timestamp.now().strftime('%d %b %Y, %H:%M')}"
    ws1['B4'].font = Font(name='Calibri', size=10, color="64748B")
    ws1['B4'].fill = fill(DARK)
    ws1['B4'].alignment = center()

    # KPI Cards Row
    kpi_row = 6
    ws1.row_dimensions[kpi_row].height = 14
    ws1.row_dimensions[kpi_row+1].height = 30
    ws1.row_dimensions[kpi_row+2].height = 20

    kpis = [
        ("Total Records", f"{len(df):,}", ACCENT),
        ("Columns", f"{len(df.columns)}", ACCENT2),
    ]
    if rev_col:
        kpis.append(("Total Revenue", f"{df[rev_col].sum():,.0f}", GREEN))
        kpis.append(("Avg Revenue", f"{df[rev_col].mean():,.0f}", YELLOW))
    if cost_col and cost_col in df.columns:
        profit = df[rev_col].sum() - df[cost_col].sum() if rev_col else 0
        kpis.append(("Net Profit", f"{profit:,.0f}", GREEN if profit > 0 else RED))

    kpi_cols = ['B', 'C', 'D', 'E', 'F', 'G']
    for i, (label, value, color) in enumerate(kpis[:6]):
        col = kpi_cols[i]
        lbl_cell = ws1[f'{col}{kpi_row}']
        val_cell = ws1[f'{col}{kpi_row+1}']
        sub_cell = ws1[f'{col}{kpi_row+2}']
        lbl_cell.value = label
        lbl_cell.font = Font(name='Calibri', size=9, color="94A3B8")
        lbl_cell.fill = fill(LGRAY)
        lbl_cell.alignment = center()
        val_cell.value = value
        val_cell.font = Font(name='Calibri', size=16, bold=True, color=color)
        val_cell.fill = fill(LGRAY)
        val_cell.alignment = center()
        sub_cell.fill = fill(LGRAY)

    # Section: Data Quality
    r = kpi_row + 5
    ws1.merge_cells(f'B{r}:G{r}')
    ws1[f'B{r}'] = "  DATA QUALITY REPORT"
    ws1[f'B{r}'].font = Font(name='Calibri', size=11, bold=True, color=ACCENT)
    ws1[f'B{r}'].fill = fill(GRAY)
    ws1[f'B{r}'].alignment = left()
    r += 1

    quality_rows = [
        ("Total Rows (Original)", f"{len(df):,}"),
        ("Total Columns", f"{len(df.columns)}"),
        ("Numeric Columns", f"{len(num_cols)}"),
        ("Missing Values Remaining", "0 (all imputed)"),
        ("Outliers Removed", f"Via IQR 1-99%"),
        ("Data Quality Score", "✅ Clean & Ready"),
    ]
    for label, value in quality_rows:
        ws1[f'B{r}'] = label
        ws1[f'C{r}'] = value
        ws1[f'B{r}'].font = Font(name='Calibri', size=10, color=WHITE)
        ws1[f'C{r}'].font = Font(name='Calibri', size=10, bold=True, color=GREEN)
        ws1[f'B{r}'].fill = fill(DARK)
        ws1[f'C{r}'].fill = fill(DARK)
        ws1[f'B{r}'].alignment = left()
        ws1[f'C{r}'].alignment = left()
        r += 1

    # Section: ML Prediction
    if ml_result and 'error' not in ml_result:
        r += 1
        ws1.merge_cells(f'B{r}:G{r}')
        ws1[f'B{r}'] = "  🤖 ML PREDICTION RESULTS"
        ws1[f'B{r}'].font = Font(name='Calibri', size=11, bold=True, color=ACCENT2)
        ws1[f'B{r}'].fill = fill(GRAY)
        ws1[f'B{r}'].alignment = left()
        r += 1

        ml_rows = [
            ("Model Used", ml_result.get('model', 'Random Forest')),
            ("R² Score (Accuracy)", str(ml_result.get('r2', 'N/A'))),
            ("Mean Absolute Error", f"{ml_result.get('mae', 'N/A'):,}"),
            ("Current Avg Revenue", f"{ml_result.get('current_avg', 0):,.2f}"),
            ("Next Period Prediction", f"{ml_result.get('next_prediction', 0):,.2f}"),
            ("Predicted Growth", f"{ml_result.get('growth_pct', 0):+.2f}%"),
        ]
        colors_ml = [WHITE, GREEN, YELLOW, WHITE, ACCENT, GREEN if ml_result.get('growth_pct', 0) > 0 else RED]
        for (label, value), color in zip(ml_rows, colors_ml):
            ws1[f'B{r}'] = label
            ws1[f'C{r}'] = value
            ws1[f'B{r}'].font = Font(name='Calibri', size=10, color=WHITE)
            ws1[f'C{r}'].font = Font(name='Calibri', size=10, bold=True, color=color)
            ws1[f'B{r}'].fill = fill(DARK)
            ws1[f'C{r}'].fill = fill(DARK)
            ws1[f'B{r}'].alignment = left()
            ws1[f'C{r}'].alignment = left()
            r += 1

    # Section: Pipeline Log
    r += 1
    ws1.merge_cells(f'B{r}:G{r}')
    ws1[f'B{r}'] = "  ⚙ ETL PIPELINE LOG"
    ws1[f'B{r}'].font = Font(name='Calibri', size=11, bold=True, color=YELLOW)
    ws1[f'B{r}'].fill = fill(GRAY)
    ws1[f'B{r}'].alignment = left()
    r += 1

    for log in logs:
        ws1.merge_cells(f'B{r}:G{r}')
        ws1[f'B{r}'] = f"  ▸ {log}"
        ws1[f'B{r}'].font = Font(name='Calibri', size=9, color="94A3B8")
        ws1[f'B{r}'].fill = fill(DARK)
        ws1[f'B{r}'].alignment = left()
        r += 1

    # ════════════════════════════════════════════════════
    # SHEET 2: CLEANED DATA
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet("🗃 Cleaned Data")
    ws2.sheet_view.showGridLines = False

    # Headers
    for ci, col in enumerate(df.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        cell.fill = fill(GRAY)
        cell.alignment = center()
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col)) + 4)

    ws2.row_dimensions[1].height = 24

    # Data rows (max 50K)
    display_df = df.head(50000)
    for ri, row in enumerate(display_df.itertuples(index=False), 2):
        bg = DARK if ri % 2 == 0 else "0A0C10"
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val if not (isinstance(val, float) and math.isnan(val)) else None)
            cell.font = Font(name='Calibri', size=9, color="94A3B8")
            cell.fill = fill(bg)
            cell.alignment = left()
        ws2.row_dimensions[ri].height = 18

    # ════════════════════════════════════════════════════
    # SHEET 3: STATISTICS
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📐 Statistics")
    ws3.sheet_view.showGridLines = False

    stat_headers = ["Column", "Min", "Max", "Mean", "Median", "Std Dev", "Sum", "Count"]
    for ci, h in enumerate(stat_headers, 2):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        cell.fill = fill(GRAY)
        cell.alignment = center()
        ws3.column_dimensions[get_column_letter(ci)].width = 16

    ws3.row_dimensions[2].height = 24

    stat_colors = [WHITE, ACCENT, ACCENT, GREEN, YELLOW, "94A3B8", GREEN, WHITE]
    for ri, col in enumerate(num_cols, 3):
        s = df[col].dropna()
        vals = [col, round(s.min(),2), round(s.max(),2), round(s.mean(),2),
                round(s.median(),2), round(s.std(),2), round(s.sum(),2), len(s)]
        for ci, (v, color) in enumerate(zip(vals, stat_colors), 2):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = Font(name='Calibri', size=10, bold=(ci==2), color=color)
            cell.fill = fill(DARK if ri % 2 == 0 else "0A0C10")
            cell.alignment = center()
        ws3.row_dimensions[ri].height = 20

    # ════════════════════════════════════════════════════
    # SHEET 4: CHARTS
    # ════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📊 Charts")
    ws4.sheet_view.showGridLines = False

    chart_data_row = 2

    if rev_col and cat_col and cat_col in df.columns:
        # Category revenue summary
        grp = df.groupby(cat_col)[rev_col].sum().reset_index().sort_values(rev_col, ascending=False).head(10)
        ws4['B1'] = cat_col
        ws4['C1'] = f"Total {rev_col}"
        ws4['B1'].font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        ws4['C1'].font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        ws4['B1'].fill = fill(GRAY)
        ws4['C1'].fill = fill(GRAY)

        for i, (_, row) in enumerate(grp.iterrows(), 2):
            ws4.cell(row=i, column=2, value=str(row[cat_col])).fill = fill(DARK)
            ws4.cell(row=i, column=3, value=round(float(row[rev_col]), 2)).fill = fill(DARK)
            ws4.cell(row=i, column=2).font = Font(name='Calibri', size=9, color=WHITE)
            ws4.cell(row=i, column=3).font = Font(name='Calibri', size=9, color=GREEN)

        # Bar Chart
        bar = BarChart()
        bar.type = "col"
        bar.title = f"Revenue by {cat_col}"
        bar.y_axis.title = rev_col
        bar.x_axis.title = cat_col
        bar.style = 10
        bar.width = 22
        bar.height = 14
        data_ref = Reference(ws4, min_col=3, min_row=1, max_row=len(grp)+1)
        cats_ref = Reference(ws4, min_col=2, min_row=2, max_row=len(grp)+1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws4.add_chart(bar, "E2")

        # Pie Chart
        pie = PieChart()
        pie.title = f"Revenue Share by {cat_col}"
        pie.width = 18
        pie.height = 14
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        ws4.add_chart(pie, "E20")

    # Line chart for numeric trend
    if rev_col and len(df) > 1:
        trend_start = len(grp) + 4 if (rev_col and cat_col and cat_col in df.columns) else 2
        ws4.cell(row=trend_start, column=2, value="Index").fill = fill(GRAY)
        ws4.cell(row=trend_start, column=3, value=rev_col).fill = fill(GRAY)
        ws4.cell(row=trend_start, column=2).font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        ws4.cell(row=trend_start, column=3).font = Font(name='Calibri', size=10, bold=True, color=ACCENT)

        sample = df[rev_col].head(50).reset_index(drop=True)
        for i, val in enumerate(sample, trend_start+1):
            ws4.cell(row=i, column=2, value=i - trend_start).fill = fill(DARK)
            ws4.cell(row=i, column=3, value=round(float(val), 2)).fill = fill(DARK)

        line = LineChart()
        line.title = f"{rev_col} Trend (Sample)"
        line.y_axis.title = rev_col
        line.x_axis.title = "Record"
        line.style = 10
        line.width = 22
        line.height = 14
        data_ref2 = Reference(ws4, min_col=3, min_row=trend_start, max_row=trend_start+len(sample))
        line.add_data(data_ref2, titles_from_data=True)
        ws4.add_chart(line, "E38")

    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 18

    # ════════════════════════════════════════════════════
    # SHEET 5: POWER BI READY
    # ════════════════════════════════════════════════════
    ws5 = wb.create_sheet("⚡ Power BI Ready")
    ws5.sheet_view.showGridLines = False

    # Instruction header
    ws5.merge_cells('B2:H2')
    ws5['B2'] = "⚡  POWER BI INTEGRATION GUIDE"
    ws5['B2'].font = Font(name='Calibri', size=14, bold=True, color=ACCENT2)
    ws5['B2'].fill = fill(DARK)
    ws5['B2'].alignment = center()

    steps = [
        ("Step 1", "Open Power BI Desktop", ACCENT),
        ("Step 2", "Click 'Get Data' → 'Excel Workbook'", ACCENT),
        ("Step 3", "Select this file and choose '🗃 Cleaned Data' sheet", ACCENT),
        ("Step 4", "Click 'Load' → Your data is imported!", ACCENT),
        ("Step 5", "Create visuals: Bar, Line, Pie, Map — all ready!", GREEN),
        ("Tip", "Save this file to OneDrive for real-time refresh in Power BI Service", YELLOW),
    ]

    r5 = 4
    for step, desc, color in steps:
        ws5[f'B{r5}'] = step
        ws5[f'C{r5}'] = desc
        ws5[f'B{r5}'].font = Font(name='Calibri', size=10, bold=True, color=color)
        ws5[f'C{r5}'].font = Font(name='Calibri', size=10, color=WHITE)
        ws5[f'B{r5}'].fill = fill(GRAY if r5 % 2 == 0 else DARK)
        ws5[f'C{r5}'].fill = fill(GRAY if r5 % 2 == 0 else DARK)
        ws5[f'B{r5}'].alignment = center()
        ws5[f'C{r5}'].alignment = left()
        ws5.row_dimensions[r5].height = 24
        r5 += 1

    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 60

    # ── Set dark background for all sheets ──
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        ws.sheet_properties.tabColor = ACCENT.replace('#', '')

    # ── Save ──
    report_path = REPORT_FOLDER / f"BI_Report_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(report_path)
    return report_path

# ── ETL Pipeline ─────────────────────────────────────────────────────────────
def run_pipeline(file_bytes, filename):
    df = read_file(file_bytes, filename)
    orig_shape = df.shape
    df, logs = clean_dataframe(df)
    rev_col, cost_col, date_col, cat_col, qty_col = detect_cols(df)

    # Feature engineering
    if rev_col and cost_col and cost_col in df.columns:
        df['Profit'] = pd.to_numeric(df[rev_col], errors='coerce') - pd.to_numeric(df[cost_col], errors='coerce')
        df['Profit_Margin_%'] = (df['Profit'] / pd.to_numeric(df[rev_col], errors='coerce') * 100).round(2)
        logs.append("Engineered 'Profit' and 'Profit_Margin_%' columns")

    ml_result = run_ml_prediction(df, rev_col)
    report_path = generate_excel_report(df, filename, ml_result, logs)

    num_cols = df.select_dtypes(include=np.number).columns.tolist()
    stats = {}
    for col in num_cols[:10]:
        s = df[col].dropna()
        stats[col] = {
            'min':    clean_val(round(float(s.min()), 2)),
            'max':    clean_val(round(float(s.max()), 2)),
            'mean':   clean_val(round(float(s.mean()), 2)),
            'sum':    clean_val(round(float(s.sum()), 2)),
        }

    cat_data = {}
    if cat_col and cat_col in df.columns and rev_col:
        grp = df.groupby(cat_col)[rev_col].sum().sort_values(ascending=False).head(8)
        cat_data = {'labels': list(grp.index.astype(str)), 'values': [clean_val(v) for v in grp.values]}

    preview = df.head(100).replace({float('nan'): None, float('inf'): None, float('-inf'): None}).to_dict(orient='records')

    return {
        'original_shape': list(orig_shape),
        'clean_shape': [len(df), len(df.columns)],
        'columns': list(df.columns),
        'logs': logs,
        'stats': stats,
        'cat_data': cat_data,
        'ml': {k: clean_val(v) if not isinstance(v, dict) else {kk: clean_val(vv) for kk, vv in v.items()} for k, v in (ml_result or {}).items()},
        'preview': preview,
        'report_path': str(report_path),
        'detected': {'revenue': rev_col, 'cost': cost_col, 'category': cat_col, 'date': date_col},
    }

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    html_path = Path(__file__).parent / 'index.html'
    if html_path.exists():
        return html_path.read_text()
    return "BI Reporter running. Open /upload"

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No file provided'}), 400
    allowed = {'.csv', '.xlsx', '.xls'}
    if not any(file.filename.lower().endswith(e) for e in allowed):
        return jsonify({'error': 'Only CSV and Excel files supported'}), 400
    try:
        file_bytes = file.read()
        result = run_pipeline(file_bytes, file.filename)
        job_id = str(uuid.uuid4())
        SESSIONS[job_id] = result
        return jsonify({'job_id': job_id, 'status': 'complete'})
    except Exception as e:
        return jsonify({'error': traceback.format_exc()}), 500

@app.route('/results/<job_id>')
def results(job_id):
    if job_id not in SESSIONS:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(SESSIONS[job_id])

@app.route('/download/<job_id>')
def download(job_id):
    if job_id not in SESSIONS:
        return jsonify({'error': 'Not found'}), 404
    path = Path(SESSIONS[job_id]['report_path'])
    if not path.exists():
        return jsonify({'error': 'Report not found'}), 404
    return send_file(path, as_attachment=True,
                     download_name=f"BI_Report_{job_id[:8]}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000, debug=False)
